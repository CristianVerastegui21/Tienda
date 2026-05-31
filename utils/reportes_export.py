import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
  Paragraph,
  SimpleDocTemplate,
  Spacer,
  Table,
  TableStyle,
)


def _estilo_tabla_pdf():
  return TableStyle([
    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
    ('FONTSIZE', (0, 0), (-1, 0), 9),
    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
    ('FONTSIZE', (0, 1), (-1, -1), 8),
    ('TOPPADDING', (0, 0), (-1, -1), 6),
    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
  ])


def generar_pdf_reporte(ruta, filtros, resumen, ventas, productos_top, stock_bajo):
  os.makedirs(os.path.dirname(ruta) or '.', exist_ok=True)

  doc = SimpleDocTemplate(
    ruta,
    pagesize=landscape(letter),
    rightMargin=36,
    leftMargin=36,
    topMargin=36,
    bottomMargin=36,
  )

  estilos = getSampleStyleSheet()
  titulo = ParagraphStyle(
    'titulo',
    parent=estilos['Title'],
    fontSize=18,
    textColor=colors.HexColor('#1e3a5f'),
    spaceAfter=4,
  )
  subtitulo = ParagraphStyle(
    'sub',
    parent=estilos['Normal'],
    fontSize=10,
    textColor=colors.HexColor('#64748b'),
    alignment=TA_LEFT,
  )

  elementos = [
    Paragraph('BODEGA VERASTEGUI', titulo),
    Paragraph('Reporte ejecutivo de ventas e inventario', subtitulo),
    Paragraph(
      f'Periodo: <b>{filtros["etiqueta"]}</b> &nbsp;|&nbsp; '
      f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}',
      subtitulo,
    ),
    Spacer(1, 14),
  ]

  kpi_data = [[
    'Total ventas',
    'Ingresos (S/.)',
    'Unidades vendidas',
    'Alertas stock',
  ], [
    str(resumen['total_ventas']),
    f"{resumen['ingresos']:.2f}",
    str(resumen['unidades']),
    str(resumen['stock_bajo_total']),
  ]]

  kpi_table = Table(kpi_data, colWidths=[2.2 * inch] * 4)
  kpi_table.setStyle(TableStyle([
    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ('FONTSIZE', (0, 1), (-1, 1), 14),
    ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
    ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor('#0f172a')),
    ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e1')),
    ('TOPPADDING', (0, 0), (-1, -1), 8),
    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
  ]))
  elementos.append(kpi_table)
  elementos.append(Spacer(1, 16))

  elementos.append(Paragraph('Detalle de ventas del periodo', estilos['Heading2']))
  elementos.append(Spacer(1, 8))

  ventas_data = [['ID', 'Fecha', 'Total (S/.)']]
  for v in ventas:
    ventas_data.append([
      str(v['id']),
      str(v['fecha'])[:19],
      f"{float(v['total']):.2f}",
    ])
  if len(ventas_data) == 1:
    ventas_data.append(['—', 'Sin ventas en este periodo', '0.00'])

  t_ventas = Table(ventas_data, colWidths=[1 * inch, 3.2 * inch, 1.5 * inch])
  t_ventas.setStyle(_estilo_tabla_pdf())
  elementos.append(t_ventas)
  elementos.append(Spacer(1, 16))

  elementos.append(Paragraph('Top productos vendidos (periodo)', estilos['Heading2']))
  elementos.append(Spacer(1, 8))

  top_data = [['Producto', 'Unidades']]
  for p in productos_top:
    top_data.append([p['nombre'], str(p['total'])])
  if len(top_data) == 1:
    top_data.append(['Sin datos', '0'])

  t_top = Table(top_data, colWidths=[5 * inch, 1.2 * inch])
  t_top.setStyle(_estilo_tabla_pdf())
  elementos.append(t_top)
  elementos.append(Spacer(1, 12))

  elementos.append(Paragraph(
    '<para align="center"><font size="8" color="#94a3b8">'
    'Documento generado automaticamente · Sistema Bodega Verastegui</font></para>',
    estilos['Normal'],
  ))

  doc.build(elementos)


def generar_excel_reporte(ruta, filtros, resumen, ventas, productos_top, stock_bajo):
  os.makedirs(os.path.dirname(ruta) or '.', exist_ok=True)

  wb = Workbook()
  header_fill = PatternFill('solid', fgColor='1E3A5F')
  header_font = Font(bold=True, color='FFFFFF', size=11)
  title_font = Font(bold=True, size=14, color='1E3A5F')
  thin = Side(style='thin', color='CBD5E1')
  border = Border(left=thin, right=thin, top=thin, bottom=thin)

  def estilo_hoja(ws, titulo):
    ws['A1'] = 'BODEGA VERASTEGUI'
    ws['A1'].font = title_font
    ws['A2'] = titulo
    ws['A3'] = f'Periodo: {filtros["etiqueta"]}'
    ws['A4'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    for r in range(1, 5):
      ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)

  ws_res = wb.active
  ws_res.title = 'Resumen'
  estilo_hoja(ws_res, 'Resumen ejecutivo')
  ws_res.append([])
  ws_res.append(['Indicador', 'Valor'])
  for c in ['A6', 'B6']:
    ws_res[c].fill = header_fill
    ws_res[c].font = header_font
    ws_res[c].border = border
  filas = [
    ('Total de ventas', resumen['total_ventas']),
    ('Ingresos (S/.)', round(resumen['ingresos'], 2)),
    ('Unidades vendidas', resumen['unidades']),
    ('Productos con stock bajo', resumen['stock_bajo_total']),
  ]
  for nombre, valor in filas:
    ws_res.append([nombre, valor])

  ws_v = wb.create_sheet('Ventas')
  estilo_hoja(ws_v, 'Detalle de ventas')
  ws_v.append([])
  headers = ['ID Venta', 'Fecha y hora', 'Total (S/.)', 'Estado']
  ws_v.append(headers)
  row = ws_v.max_row
  for col in range(1, len(headers) + 1):
    cell = ws_v.cell(row=row, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

  for v in ventas:
    ws_v.append([
      v['id'],
      str(v['fecha']),
      float(v['total']),
      'Completada',
    ])

  ws_p = wb.create_sheet('Top productos')
  estilo_hoja(ws_p, 'Productos mas vendidos')
  ws_p.append([])
  ws_p.append(['#', 'Producto', 'Unidades'])
  row = ws_p.max_row
  for col in range(1, 4):
    cell = ws_p.cell(row=row, column=col)
    cell.fill = header_fill
    cell.font = header_font
  for i, p in enumerate(productos_top, 1):
    ws_p.append([i, p['nombre'], int(p['total'])])

  ws_s = wb.create_sheet('Stock bajo')
  estilo_hoja(ws_s, 'Alertas de inventario')
  ws_s.append([])
  ws_s.append(['Producto', 'Stock', 'Punto reorden'])
  row = ws_s.max_row
  for col in range(1, 4):
    cell = ws_s.cell(row=row, column=col)
    cell.fill = header_fill
    cell.font = header_font
  for p in stock_bajo:
    ws_s.append([p['nombre'], p['stock'], p.get('reorden', '')])

  for ws in wb.worksheets:
    for col in range(1, 8):
      ws.column_dimensions[get_column_letter(col)].width = 18

  wb.save(ruta)
